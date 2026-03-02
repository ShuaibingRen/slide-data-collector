"""Tests for the validator module."""

import os
import tempfile
from pathlib import Path

import pandas as pd
import pytest

# Ensure collector package can be imported
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from collector.schema import load_schema, get_field_names, get_required_fields, get_enum_values
from collector.excel import generate_template
from collector.validator import validate_manifest, ValidationError


SCHEMAS_DIR = str(Path(__file__).parent.parent / "schemas")


class TestSchemaLoader:
    """Tests for schema loading."""

    def test_load_donor_schema(self):
        schema = load_schema(os.path.join(SCHEMAS_DIR, "donor.yaml"))
        assert schema["name"] == "Donor Metadata"
        assert len(schema["fields"]) > 0

    def test_load_sample_schema(self):
        schema = load_schema(os.path.join(SCHEMAS_DIR, "sample.yaml"))
        assert schema["name"] == "Sample Metadata"
        assert len(schema["fields"]) > 0

    def test_get_field_names(self):
        schema = load_schema(os.path.join(SCHEMAS_DIR, "donor.yaml"))
        names = get_field_names(schema)
        assert "donor_id" in names
        assert "patient_id" in names
        assert "tissue_type" in names

    def test_get_required_fields(self):
        schema = load_schema(os.path.join(SCHEMAS_DIR, "donor.yaml"))
        required = get_required_fields(schema)
        assert "donor_id" in required
        assert "notes" not in required

    def test_get_enum_values(self):
        schema = load_schema(os.path.join(SCHEMAS_DIR, "donor.yaml"))
        values = get_enum_values(schema, "tissue_type")
        assert values is not None
        assert "Lung" in values
        assert get_enum_values(schema, "donor_id") is None


class TestExcelGeneration:
    """Tests for Excel template generation."""

    def test_generate_donor_template(self, tmp_path):
        schema_path = os.path.join(SCHEMAS_DIR, "donor.yaml")
        output_path = str(tmp_path / "donor_manifest.xlsx")
        result = generate_template(schema_path, output_path)
        assert os.path.exists(result)

        # Verify the generated file has correct sheets
        xl = pd.ExcelFile(result)
        assert "Instructions - 填写说明" in xl.sheet_names
        assert "Data - 数据" in xl.sheet_names

    def test_generate_all_types(self, tmp_path):
        from collector.excel import generate_all_templates
        files = generate_all_templates(SCHEMAS_DIR, str(tmp_path))
        assert len(files) == 3
        for f in files:
            assert os.path.exists(f)


class TestValidator:
    """Tests for manifest validation."""

    def _create_manifest(self, tmp_path, data: dict, schema_type: str) -> str:
        """Helper to create a test manifest Excel file."""
        schema_path = os.path.join(SCHEMAS_DIR, f"{schema_type}.yaml")
        template_path = str(tmp_path / f"{schema_type}_manifest.xlsx")
        generate_template(schema_path, template_path)

        # Write test data to the Data sheet, starting at row 3
        from openpyxl import load_workbook
        wb = load_workbook(template_path)
        ws = wb["Data - 数据"]

        # Get header mapping
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val:
                headers[val] = col_idx

        # Write data at row 3
        for field_name, value in data.items():
            if field_name in headers:
                ws.cell(row=3, column=headers[field_name], value=value)

        wb.save(template_path)
        return template_path

    def test_valid_donor_manifest(self, tmp_path):
        manifest_path = self._create_manifest(tmp_path, {
            "donor_id": "DNR-001",
            "patient_id": "PAT-001",
            "tissue_type": "Lung",
            "diagnosis": "Adenocarcinoma",
            "collection_date": "2024-01-15",
        }, "donor")

        schema_path = os.path.join(SCHEMAS_DIR, "donor.yaml")
        errors = validate_manifest(manifest_path, schema_path)
        assert len(errors) == 0, f"Expected no errors, got: {[str(e) for e in errors]}"

    def test_missing_required_field(self, tmp_path):
        manifest_path = self._create_manifest(tmp_path, {
            "donor_id": "DNR-001",
            # patient_id missing
            "tissue_type": "Lung",
            "diagnosis": "Adenocarcinoma",
            "collection_date": "2024-01-15",
        }, "donor")

        schema_path = os.path.join(SCHEMAS_DIR, "donor.yaml")
        errors = validate_manifest(manifest_path, schema_path)
        assert any("patient_id" in e.field for e in errors)

    def test_invalid_enum_value(self, tmp_path):
        manifest_path = self._create_manifest(tmp_path, {
            "donor_id": "DNR-001",
            "patient_id": "PAT-001",
            "tissue_type": "InvalidType",  # Not in enum
            "diagnosis": "Adenocarcinoma",
            "collection_date": "2024-01-15",
        }, "donor")

        schema_path = os.path.join(SCHEMAS_DIR, "donor.yaml")
        errors = validate_manifest(manifest_path, schema_path)
        assert any("tissue_type" in e.field for e in errors)

    def test_invalid_date_format(self, tmp_path):
        manifest_path = self._create_manifest(tmp_path, {
            "donor_id": "DNR-001",
            "patient_id": "PAT-001",
            "tissue_type": "Lung",
            "diagnosis": "Adenocarcinoma",
            "collection_date": "Jan 15 2024",  # Wrong format
        }, "donor")

        schema_path = os.path.join(SCHEMAS_DIR, "donor.yaml")
        errors = validate_manifest(manifest_path, schema_path)
        assert any("collection_date" in e.field for e in errors)

    def test_valid_imaging_manifest(self, tmp_path):
        manifest_path = self._create_manifest(tmp_path, {
            "image_id": "IMG-001",
            "sample_id": "SPL-001",
            "s3_path": "s3://bucket/images/IMG-001.ome.tiff",
            "imaging_modality": "Brightfield",
            "file_format": "OME-TIFF",
            "imaging_date": "2024-02-01",
            "qc_status": "Pass",
        }, "imaging")

        schema_path = os.path.join(SCHEMAS_DIR, "imaging.yaml")
        errors = validate_manifest(manifest_path, schema_path)
        assert len(errors) == 0, f"Expected no errors, got: {[str(e) for e in errors]}"
