"""Excel template generator: creates .xlsx files with data validation."""

from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from collector.schema import load_schema, load_all_schemas


# Color definitions (matching HTAN style)
REQUIRED_FILL = PatternFill(start_color="B8D4E8", end_color="B8D4E8", fill_type="solid")  # blue
OPTIONAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # yellow
HEADER_FONT = Font(bold=True, size=11)
EXAMPLE_FONT = Font(italic=True, color="888888", size=10)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=11, color="333333")
BODY_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def generate_template(schema_path: str, output_path: str) -> str:
    """Generate an Excel template from a YAML schema.

    Args:
        schema_path: Path to the YAML schema file.
        output_path: Path for the output .xlsx file.

    Returns:
        The output file path.
    """
    schema = load_schema(schema_path)
    wb = Workbook()

    # --- Instructions Sheet ---
    ws_instr = wb.active
    ws_instr.title = "Instructions - 填写说明"
    _build_instructions_sheet(ws_instr, schema)

    # --- Data Sheet ---
    ws_data = wb.create_sheet("Data - 数据")
    _build_data_sheet(ws_data, schema)

    # Make Data sheet active by default
    wb.active = wb["Data - 数据"]

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def generate_all_templates(schemas_dir: str, output_dir: str) -> list:
    """Generate Excel templates for all schemas in a directory.

    Returns:
        List of generated file paths.
    """
    schemas = load_all_schemas(schemas_dir)
    generated = []
    for type_name, schema in schemas.items():
        output_path = str(Path(output_dir) / f"{type_name}_manifest.xlsx")
        generate_template(
            str(Path(schemas_dir) / f"{type_name}.yaml"),
            output_path,
        )
        generated.append(output_path)
    return generated


def _build_instructions_sheet(ws, schema: dict):
    """Build the Instructions worksheet."""
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 30

    # Title
    ws["A1"] = schema["name"]
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    ws["A2"] = schema.get("description", "")
    ws["A2"].font = BODY_FONT
    ws.merge_cells("A2:F2")

    ws["A4"] = "🔵 蓝色 = 必填 / Required    🟡 黄色 = 选填 / Optional"
    ws["A4"].font = SUBTITLE_FONT
    ws.merge_cells("A4:F4")

    # Field reference table header
    row = 6
    headers = ["Field Name / 字段名", "Type / 类型", "Required / 必填",
               "Description / 说明", "Example / 示例", "Allowed Values / 可选值"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.border = THIN_BORDER

    # Field rows
    for field in schema["fields"]:
        row += 1
        ws.cell(row=row, column=1, value=field["name"]).border = THIN_BORDER
        ws.cell(row=row, column=2, value=field["type"]).border = THIN_BORDER

        req_cell = ws.cell(row=row, column=3, value="✅" if field.get("required") else "")
        req_cell.border = THIN_BORDER
        req_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row, column=4, value=field.get("description", "")).border = THIN_BORDER
        ws.cell(row=row, column=5, value=field.get("example", "")).border = THIN_BORDER

        values_str = ", ".join(field["values"]) if field["type"] == "enum" else ""
        ws.cell(row=row, column=6, value=values_str).border = THIN_BORDER

        # Color required rows
        fill = REQUIRED_FILL if field.get("required") else OPTIONAL_FILL
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = fill

    # Freeze top
    ws.freeze_panes = "A7"


def _build_data_sheet(ws, schema: dict):
    """Build the Data worksheet with headers, validation, and formatting."""
    fields = schema["fields"]

    # --- Row 1: Header ---
    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=field["name"])
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        fill = REQUIRED_FILL if field.get("required") else OPTIONAL_FILL
        cell.fill = fill

        # Set column width based on field name length
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(field["name"]) + 4, 15)

    # --- Row 2: Example row (greyed out) ---
    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=2, column=col_idx, value=field.get("example", ""))
        cell.font = EXAMPLE_FONT
        cell.border = THIN_BORDER

    # --- Data validation for enum fields ---
    for col_idx, field in enumerate(fields, 1):
        if field["type"] == "enum":
            col_letter = get_column_letter(col_idx)
            formula = '"' + ",".join(str(v) for v in field["values"]) + '"'
            dv = DataValidation(
                type="list",
                formula1=formula,
                allow_blank=not field.get("required", False),
                showDropDown=False,
            )
            dv.error = f"请从下拉列表中选择 / Please select from the dropdown"
            dv.errorTitle = f"无效值 / Invalid Value"
            dv.prompt = f"可选值 / Options: {', '.join(str(v) for v in field['values'])}"
            dv.promptTitle = field["name"]
            # Apply to rows 3-1002 (1000 data rows)
            dv.add(f"{col_letter}3:{col_letter}1002")
            ws.add_data_validation(dv)

        elif field["type"] == "date":
            col_letter = get_column_letter(col_idx)
            dv = DataValidation(
                type="date",
                operator="greaterThan",
                formula1="2000-01-01",
                allow_blank=not field.get("required", False),
            )
            dv.error = "请输入有效日期 (YYYY-MM-DD) / Please enter a valid date"
            dv.errorTitle = "日期格式错误 / Invalid Date"
            dv.prompt = "格式 / Format: YYYY-MM-DD"
            dv.promptTitle = field["name"]
            dv.add(f"{col_letter}3:{col_letter}1002")
            ws.add_data_validation(dv)

        elif field["type"] == "number":
            col_letter = get_column_letter(col_idx)
            dv = DataValidation(
                type="decimal",
                operator="greaterThanOrEqual",
                formula1="0",
                allow_blank=not field.get("required", False),
            )
            dv.error = "请输入数字 / Please enter a number"
            dv.errorTitle = "数值错误 / Invalid Number"
            dv.add(f"{col_letter}3:{col_letter}1002")
            ws.add_data_validation(dv)

    # Freeze header row
    ws.freeze_panes = "A2"
