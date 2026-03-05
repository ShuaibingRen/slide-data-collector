"""Report generator: merge multi-layer manifests into a delivery report."""

from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from collector.validator import _read_manifest


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
WARN_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
OK_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def generate_report(
    donor_path: str,
    sample_path: str,
    imaging_path: str,
    output_path: str,
) -> str:
    """Merge three manifests into a single delivery report.

    Links: donor --(participant_id / parent_id)--> sample --(biospecimen_id / parent_biospecimen_id)--> imaging

    Args:
        donor_path: Path to filled donor manifest.
        sample_path: Path to filled sample manifest.
        imaging_path: Path to filled imaging manifest.
        output_path: Path for the output report.

    Returns:
        The output file path.
    """
    df_donor = _read_manifest(donor_path)
    df_sample = _read_manifest(sample_path)
    df_imaging = _read_manifest(imaging_path)

    # Prefix columns to avoid collision (except join keys)
    df_donor = df_donor.rename(
        columns={c: f"donor.{c}" for c in df_donor.columns if c != "participant_id"}
    )
    df_sample = df_sample.rename(
        columns={c: f"sample.{c}" for c in df_sample.columns if c not in ("biospecimen_id", "parent_id")}
    )
    df_imaging = df_imaging.rename(
        columns={c: f"imaging.{c}" for c in df_imaging.columns if c not in ("data_file_id", "parent_biospecimen_id")}
    )

    # Merge: imaging -> sample -> donor
    merged = df_imaging.merge(df_sample, left_on="parent_biospecimen_id", right_on="biospecimen_id", how="left")
    merged = merged.merge(df_donor, left_on="parent_id", right_on="participant_id", how="left")

    # Identify linkage issues
    merged["_link_status"] = "OK"
    merged.loc[merged["parent_id"].isna(), "_link_status"] = "Missing sample link"
    merged.loc[
        merged["donor.participant_id"].isna() & merged["parent_id"].notna(),
        "_link_status",
    ] = "Missing donor link"

    # Summary stats
    summary = {
        "Total images / 总图像数": len(df_imaging),
        "Total samples / 总样本数": len(df_sample),
        "Total donors / 总供体数": len(df_donor),
        "Linked images / 已关联图像": int((merged["_link_status"] == "OK").sum()),
        "Unlinked images / 未关联图像": int((merged["_link_status"] != "OK").sum()),
    }

    # Write to Excel
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Summary sheet
        df_summary = pd.DataFrame(
            list(summary.items()), columns=["Metric / 指标", "Value / 值"]
        )
        df_summary.to_excel(writer, sheet_name="Summary - 概览", index=False)

        # Merged data sheet
        merged.to_excel(writer, sheet_name="Merged Data - 合并数据", index=False)

        # Individual sheets for reference
        df_donor.to_excel(writer, sheet_name="Donors - 供体", index=False)
        df_sample.to_excel(writer, sheet_name="Samples - 样本", index=False)
        df_imaging.to_excel(writer, sheet_name="Images - 图像", index=False)

    # Style the output
    _style_report(output_path, merged)

    return output_path


def _style_report(path: str, merged: pd.DataFrame):
    """Apply styling to the report workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(path)

    for ws in wb.worksheets:
        # Style header row
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        # Auto-width columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

        # Freeze header
        ws.freeze_panes = "A2"

    # Highlight link issues in merged sheet
    ws_merged = wb["Merged Data - 合并数据"]
    link_col = None
    for col_idx, cell in enumerate(ws_merged[1], 1):
        if cell.value == "_link_status":
            link_col = col_idx
            break

    if link_col:
        for row in ws_merged.iter_rows(min_row=2, min_col=1, max_col=ws_merged.max_column):
            status_cell = row[link_col - 1]
            fill = OK_FILL if status_cell.value == "OK" else WARN_FILL
            for cell in row:
                cell.fill = fill

    wb.save(path)


def format_report_summary(output_path: str) -> str:
    """Read and format the summary from a generated report."""
    df = pd.read_excel(output_path, sheet_name="Summary - 概览")
    lines = ["📋 Delivery Report Summary / 交付报告概览:\n"]
    for _, row in df.iterrows():
        lines.append(f"  {row.iloc[0]}: {row.iloc[1]}")
    lines.append(f"\n  📁 Report saved to: {output_path}")
    return "\n".join(lines)
